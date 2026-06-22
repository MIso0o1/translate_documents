import openpyxl
import asyncio
import httpx
from datetime import datetime
from translate_text import translate_text_with_deepl_async

def is_numeric_or_formula(value):
    if isinstance(value, (int, float)):
        return True
    elif isinstance(value, str) and value.startswith("="):
        return True
    return False

async def translate_excel_cell_async(cell, target_lang, client):
    if cell.value and not is_numeric_or_formula(cell.value):
        value_to_translate = cell.value
        if isinstance(value_to_translate, datetime):
            value_to_translate = value_to_translate.strftime("%Y-%m-%d %H:%M:%S")
        
        translated_text = await translate_text_with_deepl_async(str(value_to_translate), target_lang, client)
        if translated_text:
            cell.value = translated_text

async def translate_excel_file_async(excel_file_path, target_lang):
    workbook = openpyxl.load_workbook(excel_file_path)
    async with httpx.AsyncClient() as client:
        tasks = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    tasks.append(translate_excel_cell_async(cell, target_lang, client))
        if tasks:
            await asyncio.gather(*tasks)
    return workbook

def translate_excel_file(excel_file_path, target_lang):
    return asyncio.run(translate_excel_file_async(excel_file_path, target_lang))
